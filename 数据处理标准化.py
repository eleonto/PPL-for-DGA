import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import unicodedata
import openpyxl

# ---------------------- 1. 读取原始数据 ----------------------
df = pd.read_excel('变压器数据最新（748组）.xlsx', sheet_name='Sheet1')

# 定义特征列和分组列
feature_cols = ['H2', 'CH4', 'C2H6', 'C2H4', 'C2H2']
group_col = '故障类型'

# ---------------------- 2. 缺失值处理：同故障类型均值填充 ----------------------
df_filled = df.copy()
# 计算每个故障类型的特征均值
group_means = df_filled.groupby(group_col)[feature_cols].mean()
# 填充缺失值
for col in feature_cols:
    df_filled[col] = df_filled.apply(
        lambda row: group_means.loc[row[group_col], col] if pd.isna(row[col]) else row[col],
        axis=1
    )

# ---------------------- 3. 异常值处理：3σ准则剔除 ----------------------
df_no_outliers = df_filled.copy()
# 按故障类型分组，对每个特征应用3σ准则
for fault_type, group in df_filled.groupby(group_col):
    for col in feature_cols:
        mu = group[col].mean()
        sigma = group[col].std()
        lower_bound = mu - 3 * sigma
        upper_bound = mu + 3 * sigma
        # 剔除异常值
        df_no_outliers = df_no_outliers[
            ~((df_no_outliers[group_col] == fault_type) & 
              ((df_no_outliers[col] < lower_bound) | (df_no_outliers[col] > upper_bound)))
        ]

# ---------------------- 4. 特征标准化：Z-score标准化 ----------------------
df_standardized = df_no_outliers.copy()
# 计算训练集均值和标准差
train_mean = df_standardized[feature_cols].mean()
train_std = df_standardized[feature_cols].std()
# Z-score标准化
df_standardized[feature_cols] = (df_standardized[feature_cols] - train_mean) / train_std

# ---------------------- 5. 计算数据分布统计 ----------------------
stats_result = []
for col in feature_cols:
    col_data = df_standardized[col]
    stats = {
        '特征气体': col,
        '数量': len(col_data),
        '缺失值数量': col_data.isna().sum(),
        '平均值': round(col_data.mean(), 3),
        '标准差': round(col_data.std(), 3),
        '最小值': round(col_data.min(), 3),
        '最大值': round(col_data.max(), 3)
    }
    stats_result.append(stats)
stats_df = pd.DataFrame(stats_result).set_index('特征气体').T

# ---------------------- 6. 保存结果到Excel并美化 ----------------------
output_file = '变压器DGA数据处理结果.xlsx'
wb = Workbook()

# 第一个Sheet：处理后完整数据
ws_data = wb.active
ws_data.title = '处理后完整数据'
# 写入表头
headers = df_standardized.columns.tolist()
for col_idx, header in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', fgColor='0070C0')
    cell.alignment = Alignment(horizontal='center', vertical='center')
# 写入数据
for row_idx, row_data in enumerate(df_standardized.values, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
        if isinstance(value, (int, float)):
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.number_format = '0.000'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        if row_idx % 2 == 0:
            cell.fill = PatternFill('solid', fgColor='EBF1F8')

# 第二个Sheet：数据分布统计
ws_stats = wb.create_sheet(title='数据分布统计')
# 写入统计表头
for row_idx, index_name in enumerate(stats_df.index, 1):
    cell = ws_stats.cell(row=row_idx + 1, column=1, value=index_name)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', fgColor='0070C0')
    cell.alignment = Alignment(horizontal='center', vertical='center')
for col_idx, col_name in enumerate(stats_df.columns, 1):
    cell = ws_stats.cell(row=1, column=col_idx + 1, value=col_name)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', fgColor='0070C0')
    cell.alignment = Alignment(horizontal='center', vertical='center')
# 写入统计数据
for row_idx, row_data in enumerate(stats_df.values, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_stats.cell(row=row_idx, column=col_idx + 1, value=value)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '0.000' if isinstance(value, float) else '0'
        if row_idx % 2 == 0:
            cell.fill = PatternFill('solid', fgColor='EBF1F8')

# 自动调整列宽
def display_width(text):
    return sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(text or ''))
def auto_fit_columns(ws, min_w=8, max_w=50, padding=3):
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        w = max((display_width(c.value) for c in col_cells
                 if not isinstance(c, openpyxl.cell.cell.MergedCell) and c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(w * 1.1 + padding, max_w))
auto_fit_columns(ws_data)
auto_fit_columns(ws_stats)

# 冻结表头
ws_data.freeze_panes = 'A2'
ws_stats.freeze_panes = 'A2'

# 保存文件
wb.save(output_file)
print(f"处理完成，结果已保存到：{output_file}")